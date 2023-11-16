#!/usr/bin/env python
# coding: utf-8

# In[5]:


from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd
from pandas import ExcelWriter
import pyautogui
import time
import openpyxl
import xlsxwriter
import numpy as np
from pynput.keyboard import Controller


def access():
    browser = webdriver.Chrome()
    browser.get("https://agenciaweb.celesc.com.br/AgenciaWeb/autenticar/loginCliente.do")
    browser.maximize_window()
    return browser
    
def save(destino):
    pyautogui.hotkey('ctrl', 's')
    time.sleep(2)
    FILE_NAME = destino
    keyboard = Controller()
    keyboard.type(FILE_NAME)
    time.sleep(0.25)
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.press('left')
    pyautogui.press('enter')
    time.sleep(1)

def find(browser, data, uc, cnpj, senha, Destino):
    search = browser.find_element("xpath",'//*[@id="fundoPrincipalLogout"]/form/div[2]/input').send_keys(uc)
    time.sleep(0.5)
    search = browser.find_element("xpath",'//*[@id="CPJ"]').click()
    time.sleep(0.5)
    search = browser.find_element("xpath",'/html/body/div[1]/table/tbody/tr[1]/td[2]/form/div[6]/input[2]').send_keys(cnpj)
    time.sleep(0.5)
    search = browser.find_element("xpath", '//*[@id="fundoPrincipalLogout"]/form/div[8]/input[1]').click()
    time.sleep(0.5)
    try:
        search = browser.find_element("xpath",'//*[@id="fundoPrincipalLogout"]/form/div[2]/input').send_keys(senha)
        search = browser.find_element("xpath", '/html/body/div[1]/table/tbody/tr[1]/td[2]/form/div[3]/input').click()
    except NoSuchElementException:
        time.sleep(0.25)
        browser.get("https://agenciaweb.celesc.com.br/AgenciaWeb/autenticar/loginCliente.do")        
    try:
        try:
            pyautogui.scroll(-2000)
            time.sleep(1.5)
            search = browser.find_element("xpath", '/html/body/div/div/div[3]/table[2]/tbody/tr[1]/td/div[1]/div/div[3]/input').click()
            search = browser.find_element(By.PARTIAL_LINK_TEXT,'»  Histórico de Pagamento').click()
        except NoSuchElementException:    
            search = browser.find_element(By.PARTIAL_LINK_TEXT,'»  Histórico de Pagamento').click()
    except NoSuchElementException:
        time.sleep(0.25)
        browser.get("https://agenciaweb.celesc.com.br/AgenciaWeb/autenticar/loginCliente.do")      
    try:
        td = browser.find_element(By.LINK_TEXT,"{}".format(data)).click()
        save(destino)
        time.sleep(0.5)
        wb.save('Base Baixar Faturas.xlsx')
        print(cliente)
        return td
    except NoSuchElementException:
        browser.get("https://agenciaweb.celesc.com.br/AgenciaWeb/autenticar/loginCliente.do")
        time.sleep(0.25)
        print('Não foi possível acessar fatura da {}'.format(cliente))
        
print('Verifique se já mudou a data no arquivo excel!')
data = str(input('Qual a data da fatura você deseja? '))
comeco = int(input('Por qual você quer começar? 2 de diferença, Ex: Quer o numero 20 então coloca 18: '))
nome_aba='Planilha2'
caminho_origem = 'S:'
nome_planilha = '\Base Baixar Faturas.xlsx'
df=pd.read_excel(caminho_origem+nome_planilha,nome_aba)
browser = access()
wb = load_workbook('Base Baixar Faturas.xlsx')
ws = wb['Planilha2']
hoje = datetime.date.today()

for i in df.index:
    if i >= comeco:
        cliente=str(df['Cliente'][i])
        uc=int(df['UC'][i])
        cnpj=str(df['CNPJ'][i])
        senha=str(df['Senha'][i])
        destino=str(df['Destino'][i])
        find(browser, data, uc, cnpj, senha, destino)
        time.sleep(1)
        browser.get("https://agenciaweb.celesc.com.br/AgenciaWeb/autenticar/loginCliente.do")
browser.quit()


# In[ ]:





# In[8]:


get_ipython().system('pip install pynput')


# In[ ]:




